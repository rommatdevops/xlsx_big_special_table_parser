import openpyxl
from openpyxl.styles import Font, Color, colors, fills, PatternFill
import re
from decimal import Decimal
from collections import Counter


redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')   

aquaFill = PatternFill(start_color='00ffff', end_color='00ffff', fill_type='solid')
	
yellowFill = PatternFill(start_color='80ff00', end_color='80ff00', fill_type='solid')  

greenFill = PatternFill(start_color='00b359', end_color='00b359', fill_type='solid')  

violetFill = PatternFill(start_color='ff00ff', end_color='ff00ff', fill_type='solid')  

orangeFill = PatternFill(start_color='ff7f00', end_color='ff7f00', fill_type='solid') 

FIRST_LINE = 5
LAST_LINE = 19950

FILENAME = "prom.22"

def openFile():
    wb = openpyxl.load_workbook(FILENAME + '.xlsx')
    return wb

def read_notification_from_column_D():
    print("Read notification from column D")
    vhodjenia_A = {}
    vhodjenia_B = {}
    row_and_value = {}
    list = []
    zakaz_numbers_dict = {}

    for i in range(FIRST_LINE, LAST_LINE):
        row_and_value[i] = sheet.cell(row=i, column=4).value
        value_column = sheet.cell(row=i, column=4).value
        x = re.search("(№| )[0-9]{9}[^0-9]", str(value_column))
        if x is not None:
            order_number = x[0].replace("№", "").replace("/", "").replace("//", "").replace(" ", "").strip()
            zakaz_numbers_dict[i] = order_number
            if sheet.cell(row=i, column=5).value != ' ':
                vhodjenia_A[i] = [order_number]
                vhodjenia_A[i].append(sheet.cell(row=i, column=5).value)
            if sheet.cell(row=i, column=6).value != ' ':
                vhodjenia_B[i] = [order_number]
                vhodjenia_B[i].append(sheet.cell(row=i, column=6).value)

    get_unique_order_number(zakaz_numbers_dict, vhodjenia_A, vhodjenia_B)

def read_commision_line():
    print("Read commision from column D")
    vhodjenia_A = {}
    vhodjenia_B = {}
    row_and_value = {}
    list = []
    zakaz_numbers_dict = {}
    order_column_A = 0
    order_column_B = 0

    for i in range(FIRST_LINE, LAST_LINE):
        row_and_value[i] = sheet.cell(row=i, column=4).value
        value_column = sheet.cell(row=i, column=4).value
        x = re.search("Автоматично додано комісія|заказ|Примітка: Заказ2", str(value_column))
        if x is not None:
            order_number = x[0].replace("№", "").replace("/", "").replace("//", "").replace(" ", "").strip()
            zakaz_numbers_dict[i] = order_number
            
            if sheet.cell(row=i, column=5).value != ' ':
                vhodjenia_A[i] = str(sheet.cell(row=i, column=5).value)
               
             
                # vhodjenia_A[i].append(sheet.cell(row=i, column=5).value)
            if sheet.cell(row=i, column=6).value != ' ':
                vhodjenia_B[i] = str(sheet.cell(row=i, column=6).value)
                if i == 801:
                    print(801, str(sheet.cell(row=i, column=6).value))
           
                # vhodjenia_B[i].append(sheet.cell(row=i, column=6).value)

    # print(vhodjenia_B['3.15'])
    return vhodjenia_B

def read_liqpay():
    # print("Read liqpay")
    vhodjenia_A = {}
    liqpay = {}
    row_and_value = {}
    list = []
    zakaz_numbers_dict = {}

    for i in range(FIRST_LINE, LAST_LINE):
        row_and_value[i] = sheet.cell(row=i, column=4).value
        value_column = sheet.cell(row=i, column=4).value
        x = re.search("Примітка: liqpay id", str(value_column))
        if x is not None:
            # order_number = x[0].replace("№", "").replace("/", "").replace("//", "").replace(" ", "").strip()
            # zakaz_numbers_dict[i] = order_number
            
            if sheet.cell(row=i, column=6).value != ' ':
                liqpay[i] = str(sheet.cell(row=i, column=6).value)
                # vhodjenia_B[i].append(sheet.cell(row=i, column=6).value)


    # print(liqpay['3.15'])
    # print(liqpay)
    return liqpay

def get_unique_order_number(order_dict, vhodjenia_A, vhodjenia_B):
    value_counts = {}
    liqpay_list = read_liqpay()
    for value in order_dict.values():
        if value in value_counts:
            value_counts[value] += 1
        else:
            value_counts[value] = 1



    unique_values = [value for value, count in value_counts.items() if count == 1]
    unique_keys = [key for key, value in order_dict.items() if value in unique_values]

    # print(unique_keys)
    fill_row(unique_keys, redFill)

    vhodjenia_A_unique = rm_unique_from_vhodjenia(unique_keys, vhodjenia_A)
    vhodjenia_B_unique = rm_unique_from_vhodjenia(unique_keys, vhodjenia_B)

    get_povernenya(vhodjenia_A_unique, vhodjenia_B_unique)
    get_different_sum(vhodjenia_A_unique, vhodjenia_B_unique)

    unique_key_value = get_unique_value(unique_keys)
    # print(unique_values)
    # for k, v in liqpay_list.items():
    #     # if v in unique_key_value.values():
    #     print(k,v)

def get_povernenya(vhodjenia_A, vhodjenia_B):
    unique_keys = []
    for k, v in vhodjenia_A.items():
        for key, value in vhodjenia_B.items():
            if v == value:
                unique_keys.append(k)
                unique_keys.append(key)
    fill_row(unique_keys, aquaFill)

def get_unique_value(unique_keys):
    # print('getuniquevalue')
    unique_keys_values = {}
    for i in unique_keys:
       
        
        if sheet.cell(row=i, column=6).value != ' ':
            unique_keys_values[i] = str(sheet.cell(row=i, column=6).value)
    # print(unique_keys_values)
    return unique_keys_values

def get_different_sum(vhodjenia_A, vhodjenia_B):
    list_of_number_commision = []
    commision_fill_row = []
    list_commision = read_commision_line()
    list_commision_reserv = list_commision
    unique_keys = []
    unique_keys_A = []
    unique_keys_B = []
    operation_different_sun_counter = 0
    without_commision = []
    # print(list_commision)
    for k, v in vhodjenia_A.items():
        
        for key, value in vhodjenia_B.items():
            # print('v[0] is ', v[0], ' v[1] is ', v[1], ' value[0] is ', value[0], ' value[1] is ', value[1])
            if v[0] == value[0] and v[1] != value[1]:
                
                # unique_keys.append(k)
                # unique_keys.append(key)
                # print(v[0], value[0])
                diff = v[1] - value[1]
                diff_formatted = str(float("{:.2f}".format(diff)))
                # print(diff_formatted)
                list_of_number_commision.append(diff_formatted)
                # if k == 534:
                #     print(f"{k} {v} {key} {value}")
                #     print(type(list_commision[801]))
                            
                result = check_diff_value_in_commision(diff_formatted, list_commision)
                if result is True:
                    k_delete = 0
                    
                    for key_commision, value_commision in list_commision.items():

                        if (value_commision == diff_formatted or float(value_commision) == float(diff_formatted)) and key_commision > key:
                            
                            # print(key_commision, value_commision, diff_formatted)
                            commision_fill_row.append(key_commision)
                            unique_keys_A.append(k)
                            unique_keys_B.append(key)

                            if key_commision in list_commision.keys():
                                # print('float')
                                k_delete = key_commision
                            break
                    if k_delete != 0:           
                        del list_commision[k_delete]
                    
                else:
                    without_commision.append(k)
                    without_commision.append(key)
    # print(list_of_number_commision)
    print("operation SUM =====", operation_different_sun_counter)
    mylist = commision_fill_row
    # print(unique_keys_A)
    print(len(set(unique_keys_A)), len(unique_keys_B), len(commision_fill_row))
    keep = {k for (k,v) in Counter(mylist).items() if v > 1}
    print([x for x in mylist if x in keep])
    # for k, v in Counter(mylist).items():
    #     print(k, v)
    fill_row(without_commision, orangeFill)
    fill_row(unique_keys_A, greenFill)
    fill_row(unique_keys_B, greenFill)
    fill_row(commision_fill_row, violetFill)


def check_diff_value_in_commision(diff_formatted, list_commision):
    
    for key_commision, value_commision in list_commision.items():
        if value_commision == diff_formatted or float(value_commision) == float(diff_formatted):
            # print(key_commision, value_commision, diff_formatted)
            return True



def rm_unique_from_vhodjenia(unique_keys, vhodjenia):
    # print(unique_keys)
    for k in unique_keys:
        if k in list(vhodjenia):
            # print(k)
            del vhodjenia[k]
    # print(vhodjenia)
    return vhodjenia

def fill_row(unique_keys, color):
    column_list = ["A", "B", "C", "D", "E", "F" ,"G"]
    # print('fill row')
    for k in unique_keys:
        for column in column_list:
            sheet[column+str(k)].fill = color

if __name__ == '__main__':
    file = openFile()
    sheet = file['Лист 1']

    read_notification_from_column_D()

    file.save(FILENAME + "_checked.xlsx")